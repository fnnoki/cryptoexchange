import os
import uuid
import io
import asyncio
import logging
import hmac
import hashlib
import secrets
import json
import time
import requests
import base58
import urllib.request
import shutil
from datetime import datetime, timedelta
from xml.etree import ElementTree as ET
import mimetypes
mimetypes.add_type('image/webp', '.webp')
mimetypes.add_type('image/png', '.png')
mimetypes.add_type('image/svg+xml', '.svg')
from dotenv import load_dotenv
from fastapi import FastAPI, Depends, HTTPException, Request, Header, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, Float, String, DateTime, Text, text, ForeignKey
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from typing import List, Optional
import uvicorn
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ---------- simple TTL cache ----------
_cache = {}
def ttl_cache(key: str, ttl_seconds: int = 30):
    def decorator(func):
        def wrapper(*args, **kwargs):
            now = time.time()
            cached = _cache.get(key)
            if cached and now - cached["ts"] < ttl_seconds:
                return cached["val"]
            val = func(*args, **kwargs)
            _cache[key] = {"val": val, "ts": now}
            return val
        return wrapper
    return decorator

def cache_clear(key: str = None):
    if key:
        _cache.pop(key, None)
    else:
        _cache.clear()

# ---------- cached template loader ----------
_template_cache = {}
def get_template(name: str) -> str:
    path = os.path.join(BASE_DIR, "templates", name)
    mtime = os.path.getmtime(path)
    cached = _template_cache.get(name)
    if cached and cached["mtime"] == mtime:
        return cached["html"]
    with open(path, "r", encoding="utf-8") as f:
        html = f.read()
    _template_cache[name] = {"html": html, "mtime": mtime}
    return html

load_dotenv()

logger.info(f"TRON_SEED set: {bool(os.getenv('TRON_SEED'))}")
logger.info(f"TRON_PRIVATE_KEY set: {bool(os.getenv('TRON_PRIVATE_KEY'))}")
logger.info(f"TRON_ADDRESS set: {bool(os.getenv('TRON_ADDRESS'))}")
logger.info(f"SOLANA_SEED set: {bool(os.getenv('SOLANA_SEED', os.getenv('TRON_SEED')))}")

# Admin authentication
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
ADMIN_SECRET = os.getenv("ADMIN_SECRET", secrets.token_hex(32))
COMMISSION_PERCENT = 3.0
HARDCODED_COIN_PRICES = {"SOL": 170, "ETH": 3500, "ARB": 0.75, "BNB": 600}

if not ADMIN_PASSWORD:
    ADMIN_PASSWORD = secrets.token_urlsafe(8)
    logger.warning(f"ADMIN PASSWORD: {ADMIN_PASSWORD}")

def generate_admin_token(password: str) -> str:
    return hmac.new(ADMIN_SECRET.encode(), password.encode(), hashlib.sha256).hexdigest()

def verify_admin_token(token: str) -> bool:
    expected = generate_admin_token(ADMIN_PASSWORD)
    return hmac.compare_digest(token, expected)

async def require_admin(authorization: Optional[str] = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    token = authorization.replace("Bearer ", "")
    if not verify_admin_token(token):
        raise HTTPException(status_code=401, detail="Неверный токен")
    return True

CURRENCY_BANKS_CONFIG = {
    "RUB": ["Сбербанк", "Тинькофф", "Альфа-Банк", "ВТБ", "Райффайзен"],
    "USD": ["Chase Bank", "Bank of America", "Wells Fargo", "Citi", "Capital One"],
    "EUR": ["Deutsche Bank", "BNP Paribas", "Société Générale", "ING", "Unicredit"],
    "GBP": ["Barclays", "HSBC", "Lloyds", "NatWest", "Santander UK"],
    "KZT": ["Halyk Bank", "Kaspi Bank", "ForteBank", "Bank CenterCredit", "Sberbank KZ"]
}

app = FastAPI(title="Crypto Exchange API")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
static_dir = os.path.join(BASE_DIR, "static")
if os.path.isdir(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

# CORS - restrict in production
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*")
origins = [o.strip() for o in ALLOWED_ORIGINS.split(",")] if ALLOWED_ORIGINS != "*" else ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000)

DB_DIR = os.path.join(BASE_DIR, "db")
DATABASE_URL = os.getenv("DATABASE_URL", "")

if DATABASE_URL:
    _is_sqlite = False
    try:
        engine = create_engine(DATABASE_URL)
    except Exception as e:
        logger.error(f"Failed to connect to database: {e}")
        logger.error("For PostgreSQL, install: pip install psycopg2-binary")
        raise
else:
    _is_sqlite = True
    os.makedirs(DB_DIR, exist_ok=True)
    if os.getenv("RENDER"):
        try:
            DB_PATH = "/var/data/orders.db"
            os.makedirs("/var/data", exist_ok=True)
        except:
            DB_PATH = os.path.join(DB_DIR, "orders.db")
    else:
        DB_PATH = os.path.join(DB_DIR, "orders.db")
        old_path = os.path.join(BASE_DIR, "orders.db")
        if os.path.isfile(old_path) and not os.path.isfile(DB_PATH):
            shutil.copy2(old_path, DB_PATH)
            logger.info(f"Migrated DB from {old_path} to {DB_PATH}")
    DATABASE_URL = f"sqlite:///{DB_PATH}"
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

class Order(Base):
    __tablename__ = "orders"
    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(String, unique=True, index=True)
    created_at = Column(DateTime, default=datetime.now)
    amount_usdt = Column(Float)
    amount_rub = Column(Float)
    rate_at_creation = Column(Float)
    commission_percent = Column(Float, default=3.0)
    commission_amount = Column(Float, default=0.0)
    currency = Column(String)
    bank = Column(String)
    phone = Column(String)
    deposit_address = Column(String)
    status = Column(String, default="pending")
    order_type = Column(String, default="buy")
    asset_type = Column(String, default="USDT")
    wallet = Column(String, nullable=True)

Base.metadata.create_all(bind=engine)

def column_exists(conn, table, column):
    if _is_sqlite:
        result = conn.execute(text(f"PRAGMA table_info({table})"))
        return column in [row[1] for row in result.fetchall()]
    else:
        result = conn.execute(text(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = :table AND column_name = :column AND table_schema = 'public'"
        ), {"table": table, "column": column})
        return result.fetchone() is not None

# Database-agnostic migrations
try:
    with engine.connect() as conn:
        migration_set = [
            ("orders", "currency", "TEXT", "UPDATE orders SET currency = 'RUB'"),
            ("orders", "order_type", "TEXT", "UPDATE orders SET order_type = 'buy'"),
            ("orders", "amount_rub", "REAL", None),
            ("orders", "rate_at_creation", "REAL", None),
            ("orders", "commission_percent", "REAL DEFAULT 3.0", None),
            ("orders", "commission_amount", "REAL DEFAULT 0.0", None),
            ("support_tickets", "status", "TEXT DEFAULT 'pending'", None),
            ("orders", "asset_type", "TEXT DEFAULT 'USDT'", None),
            ("orders", "wallet", "TEXT", None),
        ]
        for table, col, col_type, update_sql in migration_set:
            if not column_exists(conn, table, col):
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}"))
                if update_sql:
                    conn.execute(text(update_sql))
                conn.commit()
except Exception as e:
    logger.error(f"Migration error: {e}")

# WebSocket connection manager for chat
class ConnectionManager:
    def __init__(self):
        self.active_connections: dict[int, list[WebSocket]] = {}

    async def connect(self, session_id: int, websocket: WebSocket):
        await websocket.accept()
        if session_id not in self.active_connections:
            self.active_connections[session_id] = []
        self.active_connections[session_id].append(websocket)

    def disconnect(self, session_id: int, websocket: WebSocket):
        if session_id in self.active_connections:
            if websocket in self.active_connections[session_id]:
                self.active_connections[session_id].remove(websocket)
            if not self.active_connections[session_id]:
                del self.active_connections[session_id]

    async def broadcast(self, session_id: int, message: dict):
        if session_id in self.active_connections:
            for ws in self.active_connections[session_id]:
                try:
                    await ws.send_json(message)
                except:
                    pass

manager = ConnectionManager()

# ------------------------------------------------------------
# Auto-tracking: background task for pending orders
# ------------------------------------------------------------
EVM_RPCS = {
    "ETH": "https://cloudflare-eth.com",
    "ARB": "https://arb1.arbitrum.io/rpc",
    "BNB": "https://bsc-dataseed.binance.org/",
}
TRONGRID_API = "https://api.trongrid.io"
SOLANA_RPC = os.getenv("SOLANA_RPC_URL", "https://api.mainnet-beta.solana.com")
USDT_TRC20_CONTRACT = "TR7NHqjeKQxGTCi8q8ZY4pL8otSzgjLj6t"

COINGECKO_IDS = {"SOLUSDT": "solana", "ETHUSDT": "ethereum", "ARBUSDT": "arbitrum", "BNBUSDT": "binancecoin"}

def fetch_coin_usdt(symbol: str) -> Optional[float]:
    """Fetch coin price in USDT via multi-source chain. Symbol like SOLUSDT.
    Prioritizes sources accessible from Russia."""
    dash = symbol.replace("USDT", "-USDT")
    underscore = symbol.replace("USDT", "_USDT")
    lower = symbol.lower()

    # 0. CoinGecko (free, accessible from RU)
    cg_id = COINGECKO_IDS.get(symbol)
    if cg_id:
        try:
            r = requests.get(f'https://api.coingecko.com/api/v3/simple/price?ids={cg_id}&vs_currencies=usd', timeout=8)
            d = r.json()
            price = d.get(cg_id, {}).get("usd")
            if price:
                return float(price)
        except Exception:
            pass

    # 1. HTX (Huobi) — accessible from Russia, fast
    try:
        r = requests.get(f'https://api.huobi.pro/market/detail/merged?symbol={lower}', timeout=5)
        d = r.json()
        if d.get('status') == 'ok' and 'tick' in d and 'close' in d['tick']:
            return float(d['tick']['close'])
    except Exception:
        pass

    # 2. KuCoin — accessible from Russia
    try:
        r = requests.get(f'https://api.kucoin.com/api/v1/market/orderbook/level1?symbol={dash}', timeout=5)
        d = r.json()
        if d.get('code') == '200000' and 'data' in d:
            return float(d['data']['price'])
    except Exception:
        pass

    # 4. MEXC — accessible from Russia
    try:
        r = requests.get(f'https://api.mexc.com/api/v3/ticker/price?symbol={symbol}', timeout=5)
        d = r.json()
        if 'price' in d:
            return float(d['price'])
    except Exception:
        pass

    # 5. CoinEx — accessible from Russia
    try:
        r = requests.get(f'https://api.coinex.com/v1/market/ticker?market={symbol}', timeout=5)
        d = r.json()
        if d.get('code') == 0 and 'ticker' in d and 'last' in d['ticker']:
            return float(d['ticker']['last'])
    except Exception:
        pass

    # 6. Gate.io — may work from Russia
    try:
        r = requests.get(f'https://api.gateio.ws/api/v4/spot/tickers?currency_pair={underscore}', timeout=5)
        d = r.json()
        if isinstance(d, list) and len(d) > 0 and 'last' in d[0]:
            return float(d[0]['last'])
    except Exception:
        pass

    # 7. Bitget — accessible from Russia
    try:
        r = requests.get(f'https://api.bitget.com/api/v2/spot/market/tickers?symbol={symbol}', timeout=5)
        d = r.json()
        if d.get('code') == '00000' and 'data' in d and len(d['data']) > 0:
            return float(d['data'][0]['lastPr'])
    except Exception:
        pass

    # 8. OKX — may be limited from Russia
    try:
        r = requests.get(f'https://www.okx.com/api/v5/market/ticker?instId={dash}', timeout=5)
        d = r.json()
        if d.get('code') == '0' and 'data' in d and len(d['data']) > 0:
            return float(d['data'][0]['last'])
    except Exception:
        pass

    # 9. Binance (last — most likely to block RU servers)
    try:
        r = requests.get(f'https://api.binance.com/api/v3/ticker/price?symbol={symbol}', timeout=4)
        d = r.json()
        if 'price' in d:
            return float(d['price'])
    except Exception:
        pass

    # 10. OKX alternate endpoint
    try:
        r = requests.get(f'https://www.okx.cab/api/v5/market/ticker?instId={dash}', timeout=5)
        d = r.json()
        if d.get('code') == '0' and 'data' in d and len(d['data']) > 0:
            return float(d['data'][0]['last'])
    except Exception:
        pass

    return None

def check_evm_balance_sync(address: str, asset: str) -> float:
    rpc = EVM_RPCS.get(asset)
    if not rpc:
        return 0.0
    try:
        payload = {"jsonrpc": "2.0", "method": "eth_getBalance", "params": [address, "latest"], "id": 1}
        resp = requests.post(rpc, json=payload, timeout=10)
        data = resp.json()
        if "result" in data:
            return int(data["result"], 16) / 1e18
    except Exception as e:
        logger.warning(f"EVM balance error for {asset} {address}: {e}")
    return 0.0

def check_sol_balance_sync(address: str) -> float:
    try:
        payload = {"jsonrpc": "2.0", "method": "getBalance", "params": [address], "id": 1}
        resp = requests.post(SOLANA_RPC, json=payload, timeout=10)
        data = resp.json()
        if "result" in data and "value" in data["result"]:
            return data["result"]["value"] / 1e9
    except Exception as e:
        logger.warning(f"SOL balance error for {address}: {e}")
    return 0.0

def check_tron_usdt_sync(address: str) -> float:
    try:
        url = f"{TRONGRID_API}/v1/accounts/{address}"
        resp = requests.get(url, timeout=10)
        data = resp.json()
        if data.get("data") and len(data["data"]) > 0:
            tokens = data["data"][0].get("trc20", [])
            for tok in tokens:
                if USDT_TRC20_CONTRACT in tok:
                    return int(tok[USDT_TRC20_CONTRACT]) / 1e6
    except Exception as e:
        logger.warning(f"TRON USDT balance error for {address}: {e}")
    return 0.0

async def check_pending_orders_background():
    """Run every 60s: check all pending orders and auto-verify"""
    while True:
        await asyncio.sleep(60)
        try:
            db = SessionLocal()
            pending = db.query(Order).filter(Order.status == "pending").all()
            db.close()

            async def check_one(order):
                asset = (order.asset_type or "USDT").upper()
                addr = order.deposit_address
                expected = order.amount_usdt or 0
                if not addr or expected <= 0:
                    return None
                balance = 0.0
                try:
                    if asset in EVM_RPCS:
                        balance = await asyncio.to_thread(check_evm_balance_sync, addr, asset)
                    elif asset == "SOL":
                        balance = await asyncio.to_thread(check_sol_balance_sync, addr)
                    else:
                        balance = await asyncio.to_thread(check_tron_usdt_sync, addr)
                except Exception as e:
                    logger.warning(f"Background check failed for {order.order_id}: {e}")
                if balance >= expected * 0.99:
                    return order
                return None

            results = await asyncio.gather(*[check_one(o) for o in pending])
            paid_orders = [r for r in results if r is not None]

            if paid_orders:
                db2 = SessionLocal()
                for order in paid_orders:
                    try:
                        o = db2.query(Order).filter(Order.id == order.id).first()
                        if o and o.status == "pending":
                            o.status = "paid"
                            logger.info(f"Auto-verified order {o.order_id}")
                    except Exception:
                        pass
                db2.commit()
                db2.close()
        except Exception as e:
            logger.error(f"Order check background error: {e}")

# Start background task on startup
@app.on_event("startup")
async def start_background_tasks():
    asyncio.create_task(check_pending_orders_background())

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

class OrderCreate(BaseModel):
    amount_usdt: float
    currency: str
    bank: str
    phone: str
    order_type: str = "buy"
    asset_type: str = "USDT"
    wallet: Optional[str] = None

class OrderUpdate(BaseModel):
    status: str

class SupportRequest(BaseModel):
    deposit_address: Optional[str] = None
    order_id: Optional[str] = None
    email: str
    message: str

class SupportTicket(Base):
    __tablename__ = "support_tickets"
    id = Column(Integer, primary_key=True, index=True)
    created_at = Column(DateTime, default=datetime.now)
    deposit_address = Column(String, nullable=True)
    order_id = Column(String, nullable=True)
    email = Column(String)
    message = Column(String)
    status = Column(String, default="pending")

class ChatSession(Base):
    __tablename__ = "chat_sessions"
    id = Column(Integer, primary_key=True, index=True)
    client_name = Column(String, default="Клиент")
    email = Column(String, default="")
    status = Column(String, default="active")
    unread = Column(Integer, default=0)
    ip_address = Column(String(45), default="")
    country_code = Column(String(2), default="")
    country_name = Column(String(100), default="")
    wallet = Column(String(100), default="")
    created_at = Column(DateTime, default=datetime.now)

class ChatMessage(Base):
    __tablename__ = "chat_messages"
    id = Column(Integer, primary_key=True, index=True)
    session_id = Column(Integer, ForeignKey("chat_sessions.id"))
    sender = Column(String)
    message = Column(Text)
    created_at = Column(DateTime, default=datetime.now)

class Country(Base):
    __tablename__ = "countries"
    code = Column(String(2), primary_key=True)
    name = Column(String(100))
    name_ru = Column(String(100))

class BlockedWallet(Base):
    __tablename__ = "blocked_wallets"
    id = Column(Integer, primary_key=True, index=True)
    wallet = Column(String(100), unique=True, index=True)
    reason = Column(String(200), default="")
    blocked_at = Column(DateTime, default=datetime.now)
    blocked_by = Column(String, default="")

Base.metadata.create_all(bind=engine)

# Chat sessions column migrations (table exists now)
try:
    with engine.connect() as conn:
        for col, col_type in [
            ("ip_address", "VARCHAR(45) DEFAULT ''"),
            ("country_code", "VARCHAR(2) DEFAULT ''"),
            ("country_name", "VARCHAR(100) DEFAULT ''"),
            ("wallet", "VARCHAR(100) DEFAULT ''"),
        ]:
            if not column_exists(conn, "chat_sessions", col):
                conn.execute(text(f"ALTER TABLE chat_sessions ADD COLUMN {col} {col_type}"))
                conn.commit()
except Exception as e:
    logger.error(f"Chat sessions migration error: {e}")

# Seed demo support tickets if empty
try:
    db = SessionLocal()
    if db.query(SupportTicket).count() == 0:
        demo_tickets = [
            SupportTicket(email="ivan@mail.com", message="Здравствуйте! Я отправил 50 USDT, но статус заказа не меняется уже 15 минут. Заказ #A1B2C3D4. Помогите разобраться.", status="pending"),
            SupportTicket(email="maria@yandex.ru", message="Добрый день! Не могу найти свой заказ. Ввожу ID на странице отслеживания, но пишет 'не найден'. Я перевела деньги, очень переживаю.", status="pending"),
            SupportTicket(email="alex@bk.ru", message="Хотел бы уточнить по поводу лимитов. Есть ли ограничение на сумму одной операции? Планирую обменять 5000 USDT.", status="resolved"),
            SupportTicket(email="elena@mail.ru", message="Здравствуйте! Совершил перевод на старый адрес USDT. Оплата прошла, но заказ не создавался. Можете вернуть средства?", status="pending"),
            SupportTicket(email="sergey@gmail.com", message="Подскажите, какие банки поддерживаются для вывода рублей? Интересует Сбербанк и Тинькофф. И какие минимальные суммы?", status="resolved"),
        ]
        for t in demo_tickets:
            db.add(t)
        db.commit()
        logger.info(f"Seeded {len(demo_tickets)} demo support tickets")
    db.close()
except Exception as e:
    logger.error(f"Seed error: {e}")

# Seed countries if empty
try:
    db = SessionLocal()
    if db.query(Country).count() == 0:
        countries = [
            Country(code="AD", name="Andorra", name_ru="Андорра"),
            Country(code="AE", name="United Arab Emirates", name_ru="ОАЭ"),
            Country(code="AF", name="Afghanistan", name_ru="Афганистан"),
            Country(code="AG", name="Antigua and Barbuda", name_ru="Антигуа и Барбуда"),
            Country(code="AI", name="Anguilla", name_ru="Ангилья"),
            Country(code="AL", name="Albania", name_ru="Албания"),
            Country(code="AM", name="Armenia", name_ru="Армения"),
            Country(code="AO", name="Angola", name_ru="Ангола"),
            Country(code="AR", name="Argentina", name_ru="Аргентина"),
            Country(code="AS", name="American Samoa", name_ru="Американское Самоа"),
            Country(code="AT", name="Austria", name_ru="Австрия"),
            Country(code="AU", name="Australia", name_ru="Австралия"),
            Country(code="AW", name="Aruba", name_ru="Аруба"),
            Country(code="AX", name="Aland Islands", name_ru="Аландские о-ва"),
            Country(code="AZ", name="Azerbaijan", name_ru="Азербайджан"),
            Country(code="BA", name="Bosnia and Herzegovina", name_ru="Босния и Герцеговина"),
            Country(code="BB", name="Barbados", name_ru="Барбадос"),
            Country(code="BD", name="Bangladesh", name_ru="Бангладеш"),
            Country(code="BE", name="Belgium", name_ru="Бельгия"),
            Country(code="BF", name="Burkina Faso", name_ru="Буркина-Фасо"),
            Country(code="BG", name="Bulgaria", name_ru="Болгария"),
            Country(code="BH", name="Bahrain", name_ru="Бахрейн"),
            Country(code="BI", name="Burundi", name_ru="Бурунди"),
            Country(code="BJ", name="Benin", name_ru="Бенин"),
            Country(code="BL", name="Saint Barthelemy", name_ru="Сен-Бартелеми"),
            Country(code="BM", name="Bermuda", name_ru="Бермуды"),
            Country(code="BN", name="Brunei", name_ru="Бруней"),
            Country(code="BO", name="Bolivia", name_ru="Боливия"),
            Country(code="BQ", name="Bonaire", name_ru="Бонайре"),
            Country(code="BR", name="Brazil", name_ru="Бразилия"),
            Country(code="BS", name="Bahamas", name_ru="Багамы"),
            Country(code="BT", name="Bhutan", name_ru="Бутан"),
            Country(code="BW", name="Botswana", name_ru="Ботсвана"),
            Country(code="BY", name="Belarus", name_ru="Беларусь"),
            Country(code="BZ", name="Belize", name_ru="Белиз"),
            Country(code="CA", name="Canada", name_ru="Канада"),
            Country(code="CC", name="Cocos Islands", name_ru="Кокосовые о-ва"),
            Country(code="CD", name="Congo DR", name_ru="Конго (ДРК)"),
            Country(code="CF", name="Central African Republic", name_ru="ЦАР"),
            Country(code="CG", name="Congo", name_ru="Конго"),
            Country(code="CH", name="Switzerland", name_ru="Швейцария"),
            Country(code="CI", name="Cote d'Ivoire", name_ru="Кот-д'Ивуар"),
            Country(code="CK", name="Cook Islands", name_ru="Острова Кука"),
            Country(code="CL", name="Chile", name_ru="Чили"),
            Country(code="CM", name="Cameroon", name_ru="Камерун"),
            Country(code="CN", name="China", name_ru="Китай"),
            Country(code="CO", name="Colombia", name_ru="Колумбия"),
            Country(code="CR", name="Costa Rica", name_ru="Коста-Рика"),
            Country(code="CU", name="Cuba", name_ru="Куба"),
            Country(code="CV", name="Cape Verde", name_ru="Кабо-Верде"),
            Country(code="CW", name="Curacao", name_ru="Кюрасао"),
            Country(code="CX", name="Christmas Island", name_ru="Остров Рождества"),
            Country(code="CY", name="Cyprus", name_ru="Кипр"),
            Country(code="CZ", name="Czech Republic", name_ru="Чехия"),
            Country(code="DE", name="Germany", name_ru="Германия"),
            Country(code="DJ", name="Djibouti", name_ru="Джибути"),
            Country(code="DK", name="Denmark", name_ru="Дания"),
            Country(code="DM", name="Dominica", name_ru="Доминика"),
            Country(code="DO", name="Dominican Republic", name_ru="Доминиканская Республика"),
            Country(code="DZ", name="Algeria", name_ru="Алжир"),
            Country(code="EC", name="Ecuador", name_ru="Эквадор"),
            Country(code="EE", name="Estonia", name_ru="Эстония"),
            Country(code="EG", name="Egypt", name_ru="Египет"),
            Country(code="EH", name="Western Sahara", name_ru="Западная Сахара"),
            Country(code="ER", name="Eritrea", name_ru="Эритрея"),
            Country(code="ES", name="Spain", name_ru="Испания"),
            Country(code="ET", name="Ethiopia", name_ru="Эфиопия"),
            Country(code="FI", name="Finland", name_ru="Финляндия"),
            Country(code="FJ", name="Fiji", name_ru="Фиджи"),
            Country(code="FK", name="Falkland Islands", name_ru="Фолклендские о-ва"),
            Country(code="FM", name="Micronesia", name_ru="Микронезия"),
            Country(code="FO", name="Faroe Islands", name_ru="Фарерские о-ва"),
            Country(code="FR", name="France", name_ru="Франция"),
            Country(code="GA", name="Gabon", name_ru="Габон"),
            Country(code="GB", name="United Kingdom", name_ru="Великобритания"),
            Country(code="GD", name="Grenada", name_ru="Гренада"),
            Country(code="GE", name="Georgia", name_ru="Грузия"),
            Country(code="GF", name="French Guiana", name_ru="Французская Гвиана"),
            Country(code="GG", name="Guernsey", name_ru="Гернси"),
            Country(code="GH", name="Ghana", name_ru="Гана"),
            Country(code="GI", name="Gibraltar", name_ru="Гибралтар"),
            Country(code="GL", name="Greenland", name_ru="Гренландия"),
            Country(code="GM", name="Gambia", name_ru="Гамбия"),
            Country(code="GN", name="Guinea", name_ru="Гвинея"),
            Country(code="GP", name="Guadeloupe", name_ru="Гваделупа"),
            Country(code="GQ", name="Equatorial Guinea", name_ru="Экваториальная Гвинея"),
            Country(code="GR", name="Greece", name_ru="Греция"),
            Country(code="GT", name="Guatemala", name_ru="Гватемала"),
            Country(code="GU", name="Guam", name_ru="Гуам"),
            Country(code="GW", name="Guinea-Bissau", name_ru="Гвинея-Бисау"),
            Country(code="GY", name="Guyana", name_ru="Гайана"),
            Country(code="HK", name="Hong Kong", name_ru="Гонконг"),
            Country(code="HN", name="Honduras", name_ru="Гондурас"),
            Country(code="HR", name="Croatia", name_ru="Хорватия"),
            Country(code="HT", name="Haiti", name_ru="Гаити"),
            Country(code="HU", name="Hungary", name_ru="Венгрия"),
            Country(code="ID", name="Indonesia", name_ru="Индонезия"),
            Country(code="IE", name="Ireland", name_ru="Ирландия"),
            Country(code="IL", name="Israel", name_ru="Израиль"),
            Country(code="IM", name="Isle of Man", name_ru="Остров Мэн"),
            Country(code="IN", name="India", name_ru="Индия"),
            Country(code="IO", name="British Indian Ocean Territory", name_ru="Британская территория в Индийском океане"),
            Country(code="IQ", name="Iraq", name_ru="Ирак"),
            Country(code="IR", name="Iran", name_ru="Иран"),
            Country(code="IS", name="Iceland", name_ru="Исландия"),
            Country(code="IT", name="Italy", name_ru="Италия"),
            Country(code="JE", name="Jersey", name_ru="Джерси"),
            Country(code="JM", name="Jamaica", name_ru="Ямайка"),
            Country(code="JO", name="Jordan", name_ru="Иордания"),
            Country(code="JP", name="Japan", name_ru="Япония"),
            Country(code="KE", name="Kenya", name_ru="Кения"),
            Country(code="KG", name="Kyrgyzstan", name_ru="Кыргызстан"),
            Country(code="KH", name="Cambodia", name_ru="Камбоджа"),
            Country(code="KI", name="Kiribati", name_ru="Кирибати"),
            Country(code="KM", name="Comoros", name_ru="Коморы"),
            Country(code="KN", name="Saint Kitts and Nevis", name_ru="Сент-Китс и Невис"),
            Country(code="KP", name="North Korea", name_ru="Северная Корея"),
            Country(code="KR", name="South Korea", name_ru="Южная Корея"),
            Country(code="KW", name="Kuwait", name_ru="Кувейт"),
            Country(code="KY", name="Cayman Islands", name_ru="Каймановы о-ва"),
            Country(code="KZ", name="Kazakhstan", name_ru="Казахстан"),
            Country(code="LA", name="Laos", name_ru="Лаос"),
            Country(code="LB", name="Lebanon", name_ru="Ливан"),
            Country(code="LC", name="Saint Lucia", name_ru="Сент-Люсия"),
            Country(code="LI", name="Liechtenstein", name_ru="Лихтенштейн"),
            Country(code="LK", name="Sri Lanka", name_ru="Шри-Ланка"),
            Country(code="LR", name="Liberia", name_ru="Либерия"),
            Country(code="LS", name="Lesotho", name_ru="Лесото"),
            Country(code="LT", name="Lithuania", name_ru="Литва"),
            Country(code="LU", name="Luxembourg", name_ru="Люксембург"),
            Country(code="LV", name="Latvia", name_ru="Латвия"),
            Country(code="LY", name="Libya", name_ru="Ливия"),
            Country(code="MA", name="Morocco", name_ru="Марокко"),
            Country(code="MC", name="Monaco", name_ru="Монако"),
            Country(code="MD", name="Moldova", name_ru="Молдова"),
            Country(code="ME", name="Montenegro", name_ru="Черногория"),
            Country(code="MF", name="Saint Martin", name_ru="Сен-Мартен"),
            Country(code="MG", name="Madagascar", name_ru="Мадагаскар"),
            Country(code="MH", name="Marshall Islands", name_ru="Маршалловы о-ва"),
            Country(code="MK", name="North Macedonia", name_ru="Северная Македония"),
            Country(code="ML", name="Mali", name_ru="Мали"),
            Country(code="MM", name="Myanmar", name_ru="Мьянма"),
            Country(code="MN", name="Mongolia", name_ru="Монголия"),
            Country(code="MO", name="Macao", name_ru="Макао"),
            Country(code="MP", name="Northern Mariana Islands", name_ru="Северные Марианские о-ва"),
            Country(code="MQ", name="Martinique", name_ru="Мартиника"),
            Country(code="MR", name="Mauritania", name_ru="Мавритания"),
            Country(code="MS", name="Montserrat", name_ru="Монтсеррат"),
            Country(code="MT", name="Malta", name_ru="Мальта"),
            Country(code="MU", name="Mauritius", name_ru="Маврикий"),
            Country(code="MV", name="Maldives", name_ru="Мальдивы"),
            Country(code="MW", name="Malawi", name_ru="Малави"),
            Country(code="MX", name="Mexico", name_ru="Мексика"),
            Country(code="MY", name="Malaysia", name_ru="Малайзия"),
            Country(code="MZ", name="Mozambique", name_ru="Мозамбик"),
            Country(code="NA", name="Namibia", name_ru="Намибия"),
            Country(code="NC", name="New Caledonia", name_ru="Новая Каледония"),
            Country(code="NE", name="Niger", name_ru="Нигер"),
            Country(code="NF", name="Norfolk Island", name_ru="Норфолк"),
            Country(code="NG", name="Nigeria", name_ru="Нигерия"),
            Country(code="NI", name="Nicaragua", name_ru="Никарагуа"),
            Country(code="NL", name="Netherlands", name_ru="Нидерланды"),
            Country(code="NO", name="Norway", name_ru="Норвегия"),
            Country(code="NP", name="Nepal", name_ru="Непал"),
            Country(code="NR", name="Nauru", name_ru="Науру"),
            Country(code="NU", name="Niue", name_ru="Ниуэ"),
            Country(code="NZ", name="New Zealand", name_ru="Новая Зеландия"),
            Country(code="OM", name="Oman", name_ru="Оман"),
            Country(code="PA", name="Panama", name_ru="Панама"),
            Country(code="PE", name="Peru", name_ru="Перу"),
            Country(code="PF", name="French Polynesia", name_ru="Французская Полинезия"),
            Country(code="PG", name="Papua New Guinea", name_ru="Папуа — Новая Гвинея"),
            Country(code="PH", name="Philippines", name_ru="Филиппины"),
            Country(code="PK", name="Pakistan", name_ru="Пакистан"),
            Country(code="PL", name="Poland", name_ru="Польша"),
            Country(code="PM", name="Saint Pierre and Miquelon", name_ru="Сен-Пьер и Микелон"),
            Country(code="PR", name="Puerto Rico", name_ru="Пуэрто-Рико"),
            Country(code="PS", name="Palestine", name_ru="Палестина"),
            Country(code="PT", name="Portugal", name_ru="Португалия"),
            Country(code="PW", name="Palau", name_ru="Палау"),
            Country(code="PY", name="Paraguay", name_ru="Парагвай"),
            Country(code="QA", name="Qatar", name_ru="Катар"),
            Country(code="RE", name="Reunion", name_ru="Реюньон"),
            Country(code="RO", name="Romania", name_ru="Румыния"),
            Country(code="RS", name="Serbia", name_ru="Сербия"),
            Country(code="RU", name="Russia", name_ru="Россия"),
            Country(code="RW", name="Rwanda", name_ru="Руанда"),
            Country(code="SA", name="Saudi Arabia", name_ru="Саудовская Аравия"),
            Country(code="SB", name="Solomon Islands", name_ru="Соломоновы о-ва"),
            Country(code="SC", name="Seychelles", name_ru="Сейшелы"),
            Country(code="SD", name="Sudan", name_ru="Судан"),
            Country(code="SE", name="Sweden", name_ru="Швеция"),
            Country(code="SG", name="Singapore", name_ru="Сингапур"),
            Country(code="SH", name="Saint Helena", name_ru="Остров Святой Елены"),
            Country(code="SI", name="Slovenia", name_ru="Словения"),
            Country(code="SK", name="Slovakia", name_ru="Словакия"),
            Country(code="SL", name="Sierra Leone", name_ru="Сьерра-Леоне"),
            Country(code="SM", name="San Marino", name_ru="Сан-Марино"),
            Country(code="SN", name="Senegal", name_ru="Сенегал"),
            Country(code="SO", name="Somalia", name_ru="Сомали"),
            Country(code="SR", name="Suriname", name_ru="Суринам"),
            Country(code="SS", name="South Sudan", name_ru="Южный Судан"),
            Country(code="ST", name="Sao Tome and Principe", name_ru="Сан-Томе и Принсипи"),
            Country(code="SV", name="El Salvador", name_ru="Сальвадор"),
            Country(code="SX", name="Sint Maarten", name_ru="Синт-Мартен"),
            Country(code="SY", name="Syria", name_ru="Сирия"),
            Country(code="SZ", name="Eswatini", name_ru="Эсватини"),
            Country(code="TC", name="Turks and Caicos Islands", name_ru="Тёркс и Кайкос"),
            Country(code="TD", name="Chad", name_ru="Чад"),
            Country(code="TG", name="Togo", name_ru="Того"),
            Country(code="TH", name="Thailand", name_ru="Таиланд"),
            Country(code="TJ", name="Tajikistan", name_ru="Таджикистан"),
            Country(code="TK", name="Tokelau", name_ru="Токелау"),
            Country(code="TL", name="Timor-Leste", name_ru="Восточный Тимор"),
            Country(code="TM", name="Turkmenistan", name_ru="Туркменистан"),
            Country(code="TN", name="Tunisia", name_ru="Тунис"),
            Country(code="TO", name="Tonga", name_ru="Тонга"),
            Country(code="TR", name="Turkey", name_ru="Турция"),
            Country(code="TT", name="Trinidad and Tobago", name_ru="Тринидад и Тобаго"),
            Country(code="TV", name="Tuvalu", name_ru="Тувалу"),
            Country(code="TW", name="Taiwan", name_ru="Тайвань"),
            Country(code="TZ", name="Tanzania", name_ru="Танзания"),
            Country(code="UA", name="Ukraine", name_ru="Украина"),
            Country(code="UG", name="Uganda", name_ru="Уганда"),
            Country(code="US", name="United States", name_ru="США"),
            Country(code="UY", name="Uruguay", name_ru="Уругвай"),
            Country(code="UZ", name="Uzbekistan", name_ru="Узбекистан"),
            Country(code="VA", name="Vatican City", name_ru="Ватикан"),
            Country(code="VC", name="Saint Vincent and the Grenadines", name_ru="Сент-Винсент и Гренадины"),
            Country(code="VE", name="Venezuela", name_ru="Венесуэла"),
            Country(code="VG", name="British Virgin Islands", name_ru="Британские Виргинские о-ва"),
            Country(code="VI", name="US Virgin Islands", name_ru="Американские Виргинские о-ва"),
            Country(code="VN", name="Vietnam", name_ru="Вьетнам"),
            Country(code="VU", name="Vanuatu", name_ru="Вануату"),
            Country(code="WF", name="Wallis and Futuna", name_ru="Уоллис и Футуна"),
            Country(code="WS", name="Samoa", name_ru="Самоа"),
            Country(code="XK", name="Kosovo", name_ru="Косово"),
            Country(code="YE", name="Yemen", name_ru="Йемен"),
            Country(code="YT", name="Mayotte", name_ru="Майотта"),
            Country(code="ZA", name="South Africa", name_ru="ЮАР"),
            Country(code="ZM", name="Zambia", name_ru="Замбия"),
            Country(code="ZW", name="Zimbabwe", name_ru="Зимбабве"),
        ]
        for c in countries:
            db.add(c)
        db.commit()
        logger.info(f"Seeded {len(countries)} countries")
    db.close()
except Exception as e:
    logger.error(f"Countries seed error: {e}")

# --- Public pages ---

@app.get("/", response_class=HTMLResponse)
async def payment_page(request: Request):
    return get_template("payment.html")

@app.get("/sell", response_class=HTMLResponse)
async def sell_page(request: Request):
    return get_template("sell.html")

@app.get("/about", response_class=HTMLResponse)
async def about_page(request: Request):
    return get_template("about.html")

@app.get("/rules", response_class=HTMLResponse)
async def rules_page(request: Request):
    return get_template("rules.html")

@app.get("/support", response_class=HTMLResponse)
async def support_page(request: Request):
    return get_template("support.html")

# --- Admin auth ---

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    return get_template("admin_login.html")

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    return get_template("admin.html")

@app.get("/admin/support", response_class=HTMLResponse)
async def admin_support_page(request: Request):
    return get_template("admin_support.html")

@app.get("/admin/wallets", response_class=HTMLResponse)
async def admin_wallets_page(request: Request):
    return get_template("admin_wallets.html")

@app.post("/api/admin/login")
async def admin_login(request: Request):
    try:
        body = await request.json()
        password = body.get("password", "")
        if hmac.compare_digest(password, ADMIN_PASSWORD):
            token = generate_admin_token(password)
            return {"token": token}
        return JSONResponse(status_code=401, content={"detail": "Неверный пароль"})
    except Exception:
        raise HTTPException(status_code=400, detail="Ошибка запроса")

@app.get("/api/admin/check")
async def admin_check(authorization: Optional[str] = Header(None)):
    if not authorization:
        return {"authenticated": False}
    token = authorization.replace("Bearer ", "")
    return {"authenticated": verify_admin_token(token)}

@app.get("/api/admin/wallets")
async def admin_wallets_api(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Return all generated addresses with balances and linked orders"""

    # Get all unique addresses from orders
    orders = db.query(Order).all()
    addr_map = {}
    for o in orders:
        asset = (o.asset_type or "USDT").upper()
        addr = o.deposit_address
        if not addr:
            continue
        if asset not in addr_map:
            addr_map[asset] = {}
        if addr not in addr_map[asset]:
            addr_map[asset][addr] = {"address": addr, "balance": None, "orders": []}
        addr_map[asset][addr]["orders"].append({
            "order_id": o.order_id,
            "amount_usdt": o.amount_usdt,
            "status": o.status,
            "created_at": o.created_at.isoformat() if o.created_at else ""
        })

    # Fetch balances concurrently for all addresses
    async def fetch_balance(asset: str, addr: str) -> float:
        try:
            if asset in EVM_RPCS:
                payload = {"jsonrpc": "2.0", "method": "eth_getBalance", "params": [addr, "latest"], "id": 1}
                resp = await asyncio.to_thread(requests.post, EVM_RPCS[asset], json=payload, timeout=10)
                data = resp.json()
                if "result" in data:
                    return int(data["result"], 16) / 1e18
            elif asset == "SOL":
                payload = {"jsonrpc": "2.0", "method": "getBalance", "params": [addr], "id": 1}
                resp = await asyncio.to_thread(requests.post, SOLANA_RPC, json=payload, timeout=10)
                data = resp.json()
                if "result" in data and "value" in data["result"]:
                    return data["result"]["value"] / 1e9
            elif asset == "USDT":
                url = f"{TRONGRID_API}/v1/accounts/{addr}"
                resp = await asyncio.to_thread(requests.get, url, timeout=10)
                data = resp.json()
                if data.get("data") and len(data["data"]) > 0:
                    tokens = data["data"][0].get("trc20", [])
                    for tok in tokens:
                        if USDT_TRC20_CONTRACT in tok:
                            return int(tok[USDT_TRC20_CONTRACT]) / 1e6
        except Exception as e:
            logger.warning(f"Balance fetch failed for {asset} {addr}: {e}")
        return None

    tasks = []
    task_info = []
    for asset, addrs in addr_map.items():
        for addr_info in addrs.values():
            tasks.append(fetch_balance(asset, addr_info["address"]))
            task_info.append((asset, addr_info))

    results = await asyncio.gather(*tasks)
    for (asset, addr_info), balance in zip(task_info, results):
        addr_info["balance"] = balance

    result = []
    for asset, addrs in sorted(addr_map.items()):
        total_balance = sum(a["balance"] or 0 for a in addrs.values())
        total_orders = sum(len(a["orders"]) for a in addrs.values())
        result.append({
            "asset": asset,
            "total_balance": round(total_balance, 6),
            "total_orders": total_orders,
            "addresses": sorted(addrs.values(), key=lambda x: x["orders"][0]["created_at"] if x["orders"] else "", reverse=True)
        })

    seed_info = {
        "tron_seed": mask_seed(os.getenv("TRON_SEED", "")),
        "solana_seed": mask_seed(os.getenv("SOLANA_SEED", "")),
        "eth_seed": mask_seed(os.getenv("ETH_SEED", "")),
        "tron_address": os.getenv("TRON_ADDRESS", ""),
    }

    return {"assets": result, "seeds": seed_info}

def mask_seed(seed: str) -> str:
    if not seed:
        return ""
    words = seed.split()
    if len(words) <= 2:
        return seed
    return words[0] + " " + words[1] + " " + "..." + " " + words[-1]

# --- Order tracking (public) ---

@app.get("/track", response_class=HTMLResponse)
async def track_page(request: Request):
    return get_template("track.html")

@app.get("/api/orders/{order_id}/status")
async def get_order_status(order_id: str, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.order_id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    return {
        "order_id": order.order_id,
        "status": order.status,
        "amount_usdt": order.amount_usdt,
        "amount_rub": order.amount_rub,
        "currency": order.currency,
        "created_at": order.created_at.isoformat(),
        "order_type": order.order_type,
        "asset_type": order.asset_type,
        "wallet": order.wallet
    }

@app.get("/api/orders")
async def list_orders(db: Session = Depends(get_db), _=Depends(require_admin)):
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    return [{
        "order_id": o.order_id,
        "created_at": o.created_at.isoformat(),
        "amount_usdt": o.amount_usdt,
        "amount_rub": o.amount_rub,
        "currency": o.currency,
        "bank": o.bank,
        "phone": o.phone,
        "status": o.status,
        "order_type": o.order_type,
        "asset_type": o.asset_type,
        "deposit_address": o.deposit_address,
        "rate_at_creation": o.rate_at_creation,
        "commission_amount": o.commission_amount,
        "wallet": o.wallet
    } for o in orders]

@app.get("/api/orders/export")
async def export_orders_excel(db: Session = Depends(get_db), _=Depends(require_admin)):
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ["ID", "Order ID", "Created", "Type", "Asset", "Amount USDT", "Currency", "Rate", "Commission", "Bank", "Phone", "Deposit Address", "Status"]
    ws.append(headers)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    status_fills = {
        "pending": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "paid": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "canceled": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    asset_col = 5

    for i, o in enumerate(orders, 2):
        ws.cell(row=i, column=1, value=o.id)
        ws.cell(row=i, column=2, value=o.order_id)
        ws.cell(row=i, column=3, value=o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "")
        ws.cell(row=i, column=4, value="Продажа" if o.order_type == "sell" else "Покупка")
        ws.cell(row=i, column=5, value=o.asset_type or "USDT")
        ws.cell(row=i, column=6, value=o.amount_usdt)
        ws.cell(row=i, column=7, value=o.currency or "")
        ws.cell(row=i, column=8, value=o.rate_at_creation)
        ws.cell(row=i, column=9, value=o.commission_amount)
        ws.cell(row=i, column=10, value=o.bank or "")
        ws.cell(row=i, column=11, value=o.phone or "")
        ws.cell(row=i, column=12, value=o.deposit_address or "")
        ws.cell(row=i, column=13, value=o.status)

        for col in range(1, 14):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border
            if col == asset_col:
                cell.alignment = Alignment(horizontal='center')
        if o.status in status_fills:
            cell = ws.cell(row=i, column=13)
            cell.fill = status_fills[o.status]

    for col in range(1, 14):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orders_report.xlsx"}
    )

@app.get("/api/orders/{order_id}")
async def get_order(order_id: str, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.order_id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    return {
        "order_id": order.order_id,
        "status": order.status,
        "amount_usdt": order.amount_usdt,
        "amount_rub": order.amount_rub,
        "currency": order.currency,
        "created_at": order.created_at.isoformat(),
        "order_type": order.order_type,
        "asset_type": order.asset_type,
        "deposit_address": order.deposit_address,
        "wallet": order.wallet,
        "phone": order.phone,
        "bank": order.bank
    }

@app.get("/api/orders/my/{wallet}")
async def get_my_orders(wallet: str, db: Session = Depends(get_db)):
    orders = db.query(Order).filter(Order.wallet == wallet).order_by(Order.created_at.desc()).all()
    return [{
        "order_id": o.order_id,
        "created_at": o.created_at.isoformat(),
        "amount_usdt": o.amount_usdt,
        "amount_rub": o.amount_rub,
        "currency": o.currency,
        "status": o.status,
        "order_type": o.order_type,
        "asset_type": o.asset_type,
        "deposit_address": o.deposit_address
    } for o in orders]

@app.get("/account", response_class=HTMLResponse)
async def account_page(request: Request):
    return get_template("account.html")

@app.patch("/api/orders/{order_id}")
async def update_order_status(order_id: str, update: OrderUpdate, db: Session = Depends(get_db), _=Depends(require_admin)):
    order = db.query(Order).filter(Order.order_id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    if update.status not in ["pending", "paid", "canceled"]:
        raise HTTPException(status_code=400, detail="Неверный статус")
    order.status = update.status
    db.commit()
    return {"status": "success", "new_status": order.status}

# --- Public API ---

@app.post("/api/orders")
async def create_order(order: OrderCreate, db: Session = Depends(get_db)):
    logger.info(f"Creating order: {order}")

    order_id = str(uuid.uuid4())[:8].upper()

    order_count = db.query(Order).count()
    address_index = order_count

    deposit_address = None
    evm_assets = {"ETH", "ARB", "BNB"}
    if order.asset_type == "SOL":
        try:
            from solana_wallet import create_solana_address
            deposit_address = create_solana_address(address_index)
            logger.info(f"SOL deposit address: {deposit_address}")
            if not deposit_address:
                raise ValueError("No SOL address generated")
        except Exception as e:
            logger.error(f"Error creating SOL address: {e}")
            deposit_address = f"SOL{base58.b58encode(str(uuid.uuid4()).encode())[:40]}".replace('O','o')
            logger.warning(f"Using dummy SOL address: {deposit_address}")
    elif order.asset_type in evm_assets:
        try:
            from eth_wallet import create_eth_address
            deposit_address = create_eth_address(address_index)
            logger.info(f"ETH deposit address for {order.asset_type}: {deposit_address}")
            if not deposit_address:
                raise ValueError("No ETH address generated")
        except Exception as e:
            logger.error(f"Error creating ETH address: {e}")
            deposit_address = f"0x{base58.b58encode(str(uuid.uuid4()).encode())[:40]}"
            logger.warning(f"Using dummy ETH address: {deposit_address}")
    else:
        try:
            from tron_wallet import create_trc20_address
            deposit_address = create_trc20_address(address_index)
            logger.info(f"TRC20 deposit address: {deposit_address}")
            if not deposit_address:
                raise ValueError("No TRON address generated")
        except Exception as e:
            logger.error(f"Error creating TRC20 address: {e}")
            deposit_address = f"TR{str(uuid.uuid4()).replace('-', '')[:33]}"
            logger.warning(f"Using dummy TRC20 address: {deposit_address}")

    # Fetch rate and calculate with commission
    rate = None
    amount_rub = None
    commission_amount = 0.0
    try:
        cur = order.currency.upper()

        # 1. Get USDT/fiat rate
        usdt_fiat = None
        if cur == "RUB":
            try:
                r = requests.get('https://api.rapira.net/open/market/rates_xml', headers={'Accept': 'application/xml'}, timeout=5)
                r.raise_for_status()
                root = ET.fromstring(r.content)
                for item in root.findall('item'):
                    fr = item.findtext('from')
                    to = item.findtext('to')
                    out = item.findtext('out')
                    if fr == 'USDT' and to == 'RUB' and out:
                        usdt_fiat = float(out)
                        break
            except Exception:
                logger.warning("Rapira failed in create_order")

        if not usdt_fiat:
            hardcoded_fiat = {'RUB': 75.0, 'USD': 1.0, 'EUR': 0.92, 'GBP': 0.79, 'KZT': 450.0}
            usdt_fiat = hardcoded_fiat.get(cur)

        if not usdt_fiat:
            raise ValueError(f"No rate for {cur}")

        # 2. For coins, get USDT price via multi-exchange chain
        usdt_symbols = {"SOL": "SOLUSDT", "ETH": "ETHUSDT", "ARB": "ARBUSDT", "BNB": "BNBUSDT"}
        if order.asset_type in usdt_symbols:
            symbol = usdt_symbols[order.asset_type]
            coin_usdt = fetch_coin_usdt(symbol)
            if not coin_usdt:
                coin_usdt = HARDCODED_COIN_PRICES.get(order.asset_type)

            if coin_usdt:
                rate = round(coin_usdt * usdt_fiat, 6)
        else:
            rate = usdt_fiat
    except Exception as e:
        logger.warning(f"Could not fetch rate: {e}")

    if rate:
        if order.order_type == "buy":
            effective_rate = rate * (1 - COMMISSION_PERCENT / 100)
            amount_rub = round(order.amount_usdt * effective_rate, 2)
            commission_amount = round(order.amount_usdt * rate * COMMISSION_PERCENT / 100, 2)
        else:
            effective_rate = rate * (1 + COMMISSION_PERCENT / 100)
            amount_rub = round(order.amount_usdt * effective_rate, 2)
            commission_amount = round(order.amount_usdt * rate * COMMISSION_PERCENT / 100, 2)

    new_order = Order(
        order_id=order_id,
        amount_usdt=order.amount_usdt,
        amount_rub=amount_rub,
        rate_at_creation=rate,
        commission_percent=COMMISSION_PERCENT,
        commission_amount=commission_amount,
        currency=order.currency,
        bank=order.bank,
        phone=order.phone,
        deposit_address=deposit_address,
        status="pending",
        order_type=order.order_type,
        asset_type=order.asset_type,
        wallet=order.wallet
    )
    db.add(new_order)
    db.commit()
    db.refresh(new_order)

    return {
        "order_id": new_order.order_id,
        "deposit_address": new_order.deposit_address,
        "amount_usdt": new_order.amount_usdt,
        "amount_rub": new_order.amount_rub,
        "rate": rate,
        "commission_percent": new_order.commission_percent,
        "commission_amount": new_order.commission_amount,
        "status": new_order.status,
        "created_at": new_order.created_at.isoformat(),
        "asset_type": new_order.asset_type,
        "wallet": new_order.wallet
    }

@app.get("/api/currencies")
async def get_currencies():
    return {
        "currencies": list(CURRENCY_BANKS_CONFIG.keys()),
        "banks": CURRENCY_BANKS_CONFIG
    }

@app.get("/api/rate")
async def get_usdt_rate(currency: str = "RUB", asset: str = "USDT"):
    cache_key = f"rate:{currency}:{asset}"
    cached = _cache.get(cache_key)
    if cached and time.time() - cached["ts"] < 15:
        return cached["val"]
    try:
        asset_upper = asset.upper()
        cur = currency.upper()

        # 1. Get USDT/fiat rate — Rapira for RUB, hardcoded for others
        usdt_fiat = None
        if cur == "RUB":
            try:
                r = requests.get('https://api.rapira.net/open/market/rates_xml', headers={'Accept': 'application/xml'}, timeout=5)
                r.raise_for_status()
                root = ET.fromstring(r.content)
                for item in root.findall('item'):
                    fr = item.findtext('from')
                    to = item.findtext('to')
                    out = item.findtext('out')
                    if fr == 'USDT' and to == 'RUB' and out:
                        usdt_fiat = float(out)
                        break
            except Exception:
                logger.warning("Rapira failed")
        if not usdt_fiat:
            hardcoded_fiat = {'RUB': 75.0, 'USD': 1.0, 'EUR': 0.92, 'GBP': 0.79, 'KZT': 450.0}
            usdt_fiat = hardcoded_fiat.get(cur)

        if not usdt_fiat:
            result = {"error": f"Не удалось получить курс для {cur}"}
            _cache[cache_key] = {"val": result, "ts": time.time()}
            return result

        # 2. For USDT — just return fiat rate
        if asset_upper == "USDT":
            rate = usdt_fiat
            buy_rate = round(rate * (1 - COMMISSION_PERCENT / 100), 2)
            sell_rate = round(rate * (1 + COMMISSION_PERCENT / 100), 2)
            result = {"rate": rate, "buy_rate": buy_rate, "sell_rate": sell_rate, "commission_percent": COMMISSION_PERCENT, "currency": currency, "asset": "USDT", "source": "Rapira" if cur == "RUB" else "fallback"}
            _cache[cache_key] = {"val": result, "ts": time.time()}
            return result

        # 3. For coins — get USDT price via multi-exchange chain
        usdt_symbols = {"SOL": "SOLUSDT", "ETH": "ETHUSDT", "ARB": "ARBUSDT", "BNB": "BNBUSDT"}
        symbol = usdt_symbols.get(asset_upper)
        if not symbol:
            result = {"error": f"Неизвестный актив {asset_upper}"}
            _cache[cache_key] = {"val": result, "ts": time.time()}
            return result

        coin_usdt = fetch_coin_usdt(symbol)

        if not coin_usdt:
            coin_usdt = HARDCODED_COIN_PRICES.get(asset_upper)
            logger.info(f"Using hardcoded price for {asset_upper}: ${coin_usdt}")

        rate = round(coin_usdt * usdt_fiat, 6)
        buy_rate = round(rate * (1 - COMMISSION_PERCENT / 100), 2)
        sell_rate = round(rate * (1 + COMMISSION_PERCENT / 100), 2)
        result = {"rate": rate, "buy_rate": buy_rate, "sell_rate": sell_rate, "commission_percent": COMMISSION_PERCENT, "currency": currency, "asset": asset_upper, "source": "Binance+Rapira"}

        _cache[cache_key] = {"val": result, "ts": time.time()}
        return result
    except Exception as e:
        logger.error(f"Rate fetch error: {e}")
        return {"error": str(e)}

# --- Price API for chart ---

@app.get("/api/price")
async def get_asset_price(asset: str = "SOL"):
    cache_key = f"price:{asset}"
    cached = _cache.get(cache_key)
    if cached and time.time() - cached["ts"] < 20:
        return cached["val"]
    try:
        symbol_map = {"SOL": "SOLUSDT", "ETH": "ETHUSDT", "ARB": "ARBUSDT", "BNB": "BNBUSDT"}
        price = None
        asset_upper = asset.upper()

        # For USDT, return Tether's USDT/USD rate (≈1)
        if asset_upper == "USDT":
            try:
                r = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=tether&vs_currencies=usd", timeout=5)
                d = r.json()
                price = d.get("tether", {}).get("usd", 1.0)
            except Exception:
                price = 1.0
            result = {"asset": "USDT", "price": str(price), "symbol": "USDTUSD"}
            _cache[cache_key] = {"val": result, "ts": time.time()}
            return result

        symbol = symbol_map.get(asset_upper)
        if not symbol:
            return {"asset": asset_upper, "price": None, "error": "Unknown asset"}

        price = fetch_coin_usdt(symbol)

        if not price:
            price = HARDCODED_COIN_PRICES.get(asset_upper)

        if price:
            result = {"asset": asset_upper, "price": str(price), "symbol": symbol}
        else:
            result = {"asset": asset_upper, "price": None, "error": "Price unavailable"}
        _cache[cache_key] = {"val": result, "ts": time.time()}
        return result
    except Exception as e:
        logger.warning(f"Price fetch error for {asset}: {e}")
        return {"asset": asset.upper(), "price": None, "error": str(e)}

# --- Solana RPC verification ---

SOLANA_RPC_URL = os.getenv("SOLANA_RPC_URL", "https://api.mainnet-beta.solana.com")

@app.get("/api/verify_sol_tx")
async def verify_sol_tx(signature: str, expected_amount: float = 0, expected_recipient: str = ""):
    try:
        payload = {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "getTransaction",
            "params": [signature, {"encoding": "json", "maxSupportedTransactionVersion": 0}]
        }
        resp = requests.post(SOLANA_RPC_URL, json=payload, timeout=10)
        data = resp.json()

        if "result" not in data or data["result"] is None:
            return {"confirmed": False, "error": "Транзакция не найдена"}

        tx = data["result"]
        if tx.get("slot", 0) == 0:
            return {"confirmed": False}

        # Check if confirmed (has at least 1 confirmation)
        if tx.get("confirmations") is not None and tx["confirmations"] < 1:
            return {"confirmed": False}

        meta = tx.get("meta", {})
        if meta.get("err"):
            return {"confirmed": False, "error": "Транзакция не удалась"}

        return {"confirmed": True, "slot": tx.get("slot"), "blockTime": tx.get("blockTime")}
    except Exception as e:
        logger.error(f"SOL tx verification error: {e}")
        return {"confirmed": False, "error": str(e)}

@app.get("/api/check_sol_balance")
async def check_sol_balance(address: str):
    try:
        payload = {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "getBalance",
            "params": [address]
        }
        resp = requests.post(SOLANA_RPC_URL, json=payload, timeout=10)
        data = resp.json()
        if "result" not in data:
            return {"balance": 0, "error": "Не удалось получить баланс"}
        balance_lamports = data["result"]["value"]
        return {"balance": balance_lamports / 1e9, "balance_lamports": balance_lamports}
    except Exception as e:
        logger.error(f"SOL balance check error: {e}")
        return {"balance": 0, "error": str(e)}

# --- Support API ---

@app.post("/api/support")
async def create_support_ticket(req: SupportRequest, db: Session = Depends(get_db)):
    ticket = SupportTicket(
        deposit_address=req.deposit_address,
        order_id=req.order_id,
        email=req.email,
        message=req.message,
        status="pending"
    )
    db.add(ticket)
    db.commit()
    db.refresh(ticket)
    return {"id": ticket.id, "status": "pending"}

@app.get("/api/support")
async def list_support_tickets(db: Session = Depends(get_db), _=Depends(require_admin)):
    tickets = db.query(SupportTicket).order_by(SupportTicket.created_at.desc()).all()
    return [{
        "id": t.id,
        "created_at": t.created_at.isoformat(),
        "deposit_address": t.deposit_address,
        "order_id": t.order_id,
        "email": t.email,
        "message": t.message,
        "status": t.status
    } for t in tickets]

class TicketUpdate(BaseModel):
    status: str

@app.patch("/api/support/{ticket_id}")
async def update_support_ticket(ticket_id: int, update: TicketUpdate, db: Session = Depends(get_db), _=Depends(require_admin)):
    ticket = db.query(SupportTicket).filter(SupportTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    if update.status not in ["pending", "resolved", "rejected"]:
        raise HTTPException(status_code=400, detail="Неверный статус")
    ticket.status = update.status
    db.commit()
    return {"id": ticket.id, "status": ticket.status}

# --- Chat API ---

@app.post("/api/chat/create")
async def create_chat_session(request: Request, db: Session = Depends(get_db)):
    try:
        body = await request.json()
        wallet = body.get("wallet", "")

        # Check if wallet is blocked
        if wallet:
            blocked = db.query(BlockedWallet).filter(BlockedWallet.wallet == wallet).first()
            if blocked:
                raise HTTPException(status_code=403, detail="Ваш кошелек заблокирован. " + (blocked.reason or ""))

        # Get real client IP from X-Forwarded-For (Render proxy) or fallback
        forwarded = request.headers.get("x-forwarded-for", "")
        client_ip = forwarded.split(",")[0].strip() if forwarded else (request.client.host if request.client else "")

        # Try to get country from IP
        country_code = ""
        country_name = ""
        if client_ip and client_ip not in ("127.0.0.1", "::1") and not client_ip.startswith(("192.168.", "10.")):
            try:
                geo_url = f"http://ip-api.com/json/{client_ip}?fields=status,country,countryCode"
                with urllib.request.urlopen(geo_url, timeout=3) as resp:
                    geo = json.loads(resp.read())
                if geo.get("status") == "success":
                    country_code = geo.get("countryCode", "")
                    country_name = geo.get("country", "")
            except Exception:
                pass

        session = ChatSession(
            client_name=body.get("name", "Клиент"),
            email=body.get("email", ""),
            ip_address=client_ip,
            country_code=country_code,
            country_name=country_name,
            wallet=wallet,
            status="active"
        )
        db.add(session)
        db.commit()
        db.refresh(session)

        # Generate unique name for anonymous clients
        if session.client_name == "Клиент":
            session.client_name = f"Клиент #{session.id}"
            db.commit()

        # Add first message from client
        msg_text = body.get("message", "")
        if msg_text:
            msg = ChatMessage(session_id=session.id, sender="client", message=msg_text)
            db.add(msg)
            db.commit()

        return {"session_id": session.id, "status": "active"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Chat create error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/chat/sessions")
async def get_chat_sessions(db: Session = Depends(get_db), _=Depends(require_admin)):
    sessions = db.query(ChatSession).order_by(ChatSession.created_at.desc()).all()
    session_ids = [s.id for s in sessions]

    # Bulk fetch counts and last messages
    from sqlalchemy import func
    counts = dict(db.query(ChatMessage.session_id, func.count(ChatMessage.id)).filter(ChatMessage.session_id.in_(session_ids)).group_by(ChatMessage.session_id).all())
    last_msg_sub = db.query(ChatMessage.session_id, func.max(ChatMessage.id).label("max_id")).filter(ChatMessage.session_id.in_(session_ids)).group_by(ChatMessage.session_id).subquery()
    last_msgs = db.query(ChatMessage).join(last_msg_sub, ChatMessage.id == last_msg_sub.c.max_id).all()
    last_msg_map = {m.session_id: m for m in last_msgs}

    result = []
    for s in sessions:
        msg_count = counts.get(s.id, 0)
        last_msg = last_msg_map.get(s.id)
        result.append({
            "id": s.id,
            "client_name": s.client_name,
            "email": s.email,
            "status": s.status,
            "unread": s.unread,
            "ip_address": s.ip_address or "",
            "country_code": s.country_code or "",
            "country_name": s.country_name or "",
            "wallet": s.wallet or "",
            "messages_count": msg_count,
            "last_message": last_msg.message[:80] if last_msg else "",
            "last_message_time": last_msg.created_at.isoformat() if last_msg else s.created_at.isoformat(),
            "created_at": s.created_at.isoformat()
        })
    return result

@app.get("/api/chat/messages/{session_id}")
async def get_chat_messages(session_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    msgs = db.query(ChatMessage).filter(ChatMessage.session_id == session_id).order_by(ChatMessage.created_at).all()
    return [{
        "id": m.id,
        "sender": m.sender,
        "message": m.message,
        "created_at": m.created_at.isoformat()
    } for m in msgs]

@app.patch("/api/chat/close/{session_id}")
@app.post("/api/chat/close/{session_id}")
async def close_chat_session(session_id: int, request: Request, db: Session = Depends(get_db)):
    auth = request.headers.get("authorization", "")
    if not auth or not verify_admin_token(auth.replace("Bearer ", "")):
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Сессия не найдена")
    session.status = "closed"
    db.commit()
    await manager.broadcast(session_id, {"type": "closed"})
    logger.info(f"Chat session {session_id} closed by admin")
    return {"status": "closed"}

@app.post("/api/chat/block")
async def block_wallet(request: Request, db: Session = Depends(get_db)):
    try:
        auth = request.headers.get("authorization", "")
        if not auth or not verify_admin_token(auth.replace("Bearer ", "")):
            raise HTTPException(status_code=401, detail="Требуется авторизация")
        body = await request.json()
        wallet = body.get("wallet", "").strip()
        reason = body.get("reason", "").strip()
        if not wallet:
            raise HTTPException(status_code=400, detail="Укажите кошелек")
        existing = db.query(BlockedWallet).filter(BlockedWallet.wallet == wallet).first()
        if existing:
            return {"status": "already_blocked"}
        blocked = BlockedWallet(wallet=wallet, reason=reason, blocked_by="admin")
        db.add(blocked)
        db.commit()
        # Close all active sessions from this wallet
        sessions = db.query(ChatSession).filter(ChatSession.wallet == wallet, ChatSession.status == "active").all()
        for s in sessions:
            s.status = "closed"
            await manager.broadcast(s.id, {"type": "closed"})
        db.commit()
        logger.info(f"Wallet {wallet} blocked by admin")
        return {"status": "blocked"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Block error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/chat/unblock")
async def unblock_wallet(request: Request, db: Session = Depends(get_db)):
    try:
        auth = request.headers.get("authorization", "")
        if not auth or not verify_admin_token(auth.replace("Bearer ", "")):
            raise HTTPException(status_code=401, detail="Требуется авторизация")
        body = await request.json()
        wallet = body.get("wallet", "").strip()
        if not wallet:
            raise HTTPException(status_code=400, detail="Укажите кошелек")
        blocked = db.query(BlockedWallet).filter(BlockedWallet.wallet == wallet).first()
        if blocked:
            db.delete(blocked)
            db.commit()
        logger.info(f"Wallet {wallet} unblocked by admin")
        return {"status": "unblocked"}
    except Exception as e:
        logger.error(f"Unblock error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/chat/blocked")
async def get_blocked_wallets(db: Session = Depends(get_db), _=Depends(require_admin)):
    wallets = db.query(BlockedWallet).order_by(BlockedWallet.blocked_at.desc()).all()
    return [{"wallet": w.wallet, "reason": w.reason, "blocked_at": w.blocked_at.isoformat()} for w in wallets]

@app.patch("/api/chat/read/{session_id}")
async def mark_chat_read(session_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
    if session:
        session.unread = 0
        db.commit()
    return {"ok": True}

@app.post("/api/chat/send")
async def send_chat_message(request: Request, db: Session = Depends(get_db)):
    try:
        body = await request.json()
        session_id = body.get("session_id")
        message = body.get("message", "").strip()
        if not session_id or not message:
            raise HTTPException(status_code=400, detail="Missing session_id or message")
        session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        auth = request.headers.get("authorization", "")
        is_admin = False
        if auth:
            token = auth.replace("Bearer ", "")
            is_admin = verify_admin_token(token)
        sender = "admin" if is_admin else "client"
        msg = ChatMessage(session_id=session_id, sender=sender, message=message)
        db.add(msg)
        db.commit()
        db.refresh(msg)
        if not is_admin:
            session.unread = (session.unread or 0) + 1
            db.commit()
        await manager.broadcast(session_id, {"type": "message", "id": msg.id, "sender": sender, "message": message, "created_at": msg.created_at.isoformat(), "ip_address": session.ip_address or "", "country_code": session.country_code or "", "country_name": session.country_name or ""})
        return {"id": msg.id, "sender": sender, "message": message, "created_at": msg.created_at.isoformat()}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/chat/poll/{session_id}")
async def poll_chat_messages(session_id: int, since_id: int = 0, db: Session = Depends(get_db)):
    msgs = db.query(ChatMessage).filter(ChatMessage.session_id == session_id, ChatMessage.id > since_id).order_by(ChatMessage.created_at).all()
    return [{"id": m.id, "sender": m.sender, "message": m.message, "created_at": m.created_at.isoformat()} for m in msgs]

@app.websocket("/ws/chat/{session_id}")
async def chat_websocket(websocket: WebSocket, session_id: int, token: Optional[str] = ""):
    db = SessionLocal()
    try:
        session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
        if not session:
            await websocket.close(code=4004)
            return

        is_admin = False
        if token and verify_admin_token(token):
            is_admin = True

        await manager.connect(session_id, websocket)

        while True:
            data = await websocket.receive_text()
            msg_data = json.loads(data)
            msg_text = msg_data.get("message", "").strip()
            if not msg_text:
                continue

            sender = "admin" if is_admin else "client"
            msg = ChatMessage(session_id=session_id, sender=sender, message=msg_text)
            db.add(msg)
            db.commit()
            db.refresh(msg)

            if not is_admin:
                session.unread = (session.unread or 0) + 1
                db.commit()

            broadcast_data = {
                "type": "message",
                "id": msg.id,
                "sender": sender,
                "message": msg_text,
                "created_at": msg.created_at.isoformat(),
            }
            # Include IP/country info in broadcast (for admin display)
            if is_admin:
                # Client info doesn't change
                pass
            else:
                broadcast_data["ip_address"] = session.ip_address or ""
                broadcast_data["country_code"] = session.country_code or ""
                broadcast_data["country_name"] = session.country_name or ""
            await manager.broadcast(session_id, broadcast_data)

    except WebSocketDisconnect:
        manager.disconnect(session_id, websocket)
    except Exception as e:
        logger.error(f"WebSocket error: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=port,
        proxy_headers=True,
        forwarded_allow_ips='*',
        ws_ping_interval=25,
        ws_ping_timeout=10,
    )
